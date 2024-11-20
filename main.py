import datetime
import itertools
from binance.um_futures import UMFutures
import pandas as pd
import okx.MarketData as MarketData
from pybit.unified_trading import HTTP
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
from kraken.futures import Market


#TODO

#add time 

#sources : binance, okx , bybit, bitget, kraken, bingx, kucoin
#add : phemex,  crypto.com?, whitebit?

#notpreferred : deribit was added but not used in this version.
#PLAN, ADD A FEW MORE COMMON EXCHANGE.
#GUI MAY BE ADDED LATER. 





arb_threshold = 0.5  #  threshold for arbitrage opportunities, def value = 1 


## BINANCE DATA RETRIEVAL 
# # Initialize the client
client = UMFutures()

# Retrieve exchange information
exchange_info = client.exchange_info()
symbols = exchange_info['symbols']

# Fetch current prices
prices = client.ticker_price()
price_dict = {item['symbol']: item['price'] for item in prices}

# Display contract details with current prices
# for symbol_info in symbols:
#     symbol = symbol_info['symbol']
#     contract_type = symbol_info['contractType']
#     base_asset = symbol_info['baseAsset']
#     quote_asset = symbol_info['quoteAsset']
#     price = price_dict.get(symbol, 'N/A')
#     print(f"Symbol: {symbol}, Contract Type: {contract_type}, Base Asset: {base_asset}, Quote Asset: {quote_asset}, Current Price: {price}")

data = []
for symbol_info in symbols:
    symbol = symbol_info['symbol']
    base_asset = symbol_info['baseAsset']
    quote_asset = symbol_info['quoteAsset']
    contract_type = symbol_info['contractType']
    price = price_dict.get(symbol, 'N/A')
    data.append({
        'symbol': symbol,
        'Base Asset': base_asset,
        'Quote Asset': quote_asset,
        'Contract Type': contract_type,
        'bn_lastPrice': price # changed from 'Current Price' to 'Last Price'
    })

df_binance = pd.DataFrame(data)

# Display the DataFrame
print(df_binance)



##OKX DATA RETRIEVAL
# ##now okx


flag = "0"  # Production trading:0 , demo trading:1

marketDataAPI =  MarketData.MarketAPI(flag=flag)

# Retrieve the latest price snapshot, best bid/ask price, and trading volume in the last 24 hours
result = marketDataAPI.get_tickers(
    instType="SWAP"
)
# print(result)
# Assuming your result is stored in a variable called 'result'
data = result['data']  # Extract the list of dictionaries containing futures data

# Create a DataFrame from the data
df2 = pd.DataFrame(data)

# # Modify the 'instId' column to remove the last part (e.g., convert "BTC-USDT-SWAP" to "BTC-USDT")
# df2['symbol'] = df2['instId'].apply(lambda x: '-'.join(x.split('-')[:1]))

# Modify the 'instId' column to convert "BTC-USDT-SWAP" to "BTCUSDT"
df2['symbol'] = df2['instId'].apply(lambda x: x.replace('-SWAP', '').replace('-', ''))


# Rename the 'last' column to 'lastPrice'
df2 = df2.rename(columns={'last': 'okx_lastPrice'})

# Drop the original 'instId' column if it's no longer needed
df_okx = df2.drop(columns=['instId'])

# Display the DataFrame to ensure data is organized as desired
print(df_okx)



##BYBIT DATA RETRIEVAL
##
session = HTTP(testnet=False)
response = session.get_instruments_info(category='linear')
contracts = response['result']['list']
usdt_contracts = [contract for contract in contracts 
                  if contract['quoteCoin'] == 'USDT']
tickers = session.get_tickers(category='linear')['result']['list']
# Create a DataFrame for contract information
df_contracts = pd.DataFrame(usdt_contracts)

# Create a DataFrame for ticker information
df_tickers = pd.DataFrame(tickers)

# Merge the two DataFrames on the 'symbol' column
df3 = pd.merge(df_contracts, df_tickers, on='symbol', suffixes=('_contract', '_ticker'))

# Select relevant columns
df_bybit = df3[['symbol', 'baseCoin', 'quoteCoin', 'lastPrice', 'highPrice24h', 'lowPrice24h', 'volume24h']]
df_bybit = df_bybit.rename(columns={'lastPrice': 'bybit_lastPrice'})

print(df_bybit)


#
##BITGET DATA RETRIEVAL 

# Bitget public API URL for market data (USDT-margined perpetual futures)
url = 'https://api.bitget.com/api/mix/v1/market/tickers?productType=umcbl'

# Make the GET request
response = requests.get(url)
data = response.json()

# Extract relevant data
if 'data' in data:
    market_data = data['data']
    # Create a DataFrame
    df4 = pd.DataFrame(market_data)
    # Select and rename relevant columns
    df4 = df4[['symbol', 'last']]
    df4.columns = ['symbol', 'bitget_lastPrice']
    # Assuming your Bitget DataFrame is named df with a 'Symbol' column
    df4['symbol'] = df4['symbol'].apply(lambda x: x.split('_')[0])
    # Display the DataFrame
    print(df4)
else:
    print("No data received from Bitget API.")


#
##EXCH 2 DATA RETRIEVAL  KRAKEN, TOTAL OF 5.
import requests
import pandas as pd

# Kraken Futures API URL for tickers
url = 'https://futures.kraken.com/derivatives/api/v3/tickers'

# Fetch the data
response = requests.get(url)
data = response.json()

# Ensure the response is successful and contains the expected data
if data['result'] == 'success' and 'tickers' in data:
    tickers = data['tickers']

    # Filter for perpetual contracts
    perpetuals = [ticker for ticker in tickers if ticker.get('tag') == 'perpetual']

    # Extract relevant data into a structured format
    data_list = []
    for ticker in perpetuals:
        original_symbol = ticker['symbol']
        modified_symbol = original_symbol.replace('PF_', '')  # Remove 'PF_'
        modified_symbol = modified_symbol.replace('USD', 'USDT')  # Replace 'USD' with 'USDT'

        data_list.append({
            'symbol': modified_symbol,
            'kr_lastPrice': ticker['last'],
            # 'Bid Price': ticker['bid'],
            # 'Ask Price': ticker['ask'],
            '24h Volume': ticker['vol24h'],
            'Open Interest': ticker['openInterest'],
            # '24h High': ticker['high24h'],
            # '24h Low': ticker['low24h']
        })

    # Create a DataFrame
    df_kraken = pd.DataFrame(data_list)

    # Display the DataFrame
    print("kraken KRAKEN kraken")
    print(df_kraken)

    # Optionally, save to an Excel file
    # df_kraken.to_excel('kraken_perpetual_prices.xlsx', index=False)
else:
    print("Failed to retrieve data or no tickers found.")


#
## EXCH 3 DATA RETRIEVAL, TOTAL OF 6. BINGX, TOTAL OF 6.

# BingX API endpoint for futures price data
url = 'https://open-api.bingx.com/openApi/swap/v1/ticker/price'

# Fetch the data
response = requests.get(url)
data = response.json()

# Check if the response is successful
if data['code'] == 0:
    contracts = data['data']

    # Extract relevant data into a structured format
    data_list = []
    for contract in contracts:
        # Extract symbol and modify it to match your existing data format
        original_symbol = contract['symbol']  # e.g., '10000000AIDOGE-USDT'
        # Standardize symbol by removing numbers and modifying to match existing format (e.g., 'AIDOGEUSDT')
        modified_symbol = ''.join([char for char in original_symbol if not char.isdigit()]).replace('-', '')

        data_list.append({
            'symbol': modified_symbol,
            'bingx_lastPrice': float(contract['price']),
            'Time': pd.to_datetime(contract['time'], unit='ms')  # Convert timestamp to readable format
        })

    # Create a DataFrame
    df_bingx = pd.DataFrame(data_list)

    # Display the DataFrame
    print(df_bingx)

    # Optionally, save to an Excel file
    # df_bingx.to_excel('bingx_perpetual_prices.xlsx', index=False)
else:
    print("Failed to retrieve data from BingX API.")

#
##EXCH 4 DATA RETRIEVAL, TOTAL OF 7. KUCOIN, TOTAL OF 7.

# KuCoin API endpoint for futures tickers
url = 'https://api-futures.kucoin.com/api/v1/allTickers'

# Fetch the data
response = requests.get(url)
data = response.json()

# Check if the response is successful
if data['code'] == '200000':
    tickers = data['data']

    # Extract relevant data into a structured format
    data_list = []
    for ticker in tickers:
        original_symbol = ticker['symbol']  # e.g., 'SNXUSDTM'
        # Modify the symbol to remove the trailing 'M' for consistency (e.g., 'SNXUSDTM' -> 'SNXUSDT')
        modified_symbol = original_symbol.replace('M', '')

        data_list.append({
            'symbol': modified_symbol,
            'kucoin_lastPrice': float(ticker['price']),
            'Bid Price': float(ticker['bestBidPrice']),
            'Ask Price': float(ticker['bestAskPrice']),
            'Trade Size': int(ticker['size']),
            'Best Bid Size': int(ticker['bestBidSize']),
            'Best Ask Size': int(ticker['bestAskSize']),
            # 'Timestamp': pd.to_datetime(ticker['ts'], unit='ms')  # Convert timestamp to readable format
            
            'Timestamp': (ticker['ts'])
        })

    # Create a DataFrame
    df_kucoin = pd.DataFrame(data_list)

    # Display the DataFrame
    print(df_kucoin)

    # Optionally, save to an Excel file
    # df_kucoin.to_excel('kucoin_futures_prices.xlsx', index=False)
else:
    print("Failed to retrieve data from KuCoin API.")

#
##ECXH 5 DATA RETRIEVAL, TOTAL OF 8. PHEMEX, TOTAL OF 8.


##
#

df_binance = df_binance[['symbol', 'bn_lastPrice']]
df_okx = df_okx[['symbol', 'okx_lastPrice']]            
df_bybit = df_bybit[['symbol', 'bybit_lastPrice']]  
df_bitget = df4[['symbol', 'bitget_lastPrice']]
df_kraken = df_kraken[['symbol', 'kr_lastPrice']]
df_bingx = df_bingx[['symbol', 'bingx_lastPrice']]
df_kucoin = df_kucoin[['symbol', 'kucoin_lastPrice']]

df_binance["bn_lastPrice"] = pd.to_numeric(df_binance["bn_lastPrice"], errors='coerce')
df_okx["okx_lastPrice"] = pd.to_numeric(df_okx["okx_lastPrice"], errors='coerce')
df_bybit["bybit_lastPrice"] = pd.to_numeric(df_bybit["bybit_lastPrice"], errors='coerce')   
df_bitget["bitget_lastPrice"] = pd.to_numeric(df_bitget["bitget_lastPrice"], errors='coerce')   
df_kraken["kr_lastPrice"] = pd.to_numeric(df_kraken["kr_lastPrice"], errors='coerce')
df_bingx["bingx_lastPrice"] = pd.to_numeric(df_bingx["bingx_lastPrice"], errors='coerce')   
df_kucoin["kucoin_lastPrice"] = pd.to_numeric(df_kucoin["kucoin_lastPrice"], errors='coerce')   

#after this step, add the EXC_lastPrice to this list: 
# price_columns = ['bn_lastPrice', 'okx_lastPrice', 'bybit_lastPrice', 'bitget_lastPrice']


merged_data = pd.merge(df_binance, df_okx, on='symbol', how='outer')
merged_data = pd.merge(merged_data, df_bybit, on='symbol', how='outer') 
merged_data = pd.merge(merged_data, df_bitget, on='symbol', how='outer')  
merged_data = pd.merge(merged_data, df_kraken, on='symbol', how='outer')
merged_data = pd.merge(merged_data, df_bingx, on='symbol', how='outer')
merged_data = pd.merge(merged_data, df_kucoin, on='symbol', how='outer')


# print(merged_data)
# Identify symbols present in two or more exchanges
merged_data['exchange_count'] = merged_data[['bn_lastPrice', 'okx_lastPrice', 'bybit_lastPrice','bitget_lastPrice', 'kr_lastPrice','bingx_lastPrice','kucoin_lastPrice']].notnull().sum(axis=1)
potential_arbitrage = merged_data[merged_data['exchange_count'] >= 2].drop(columns='exchange_count')





# # Define new columns for percentage differences between exchanges
merged_data['perc_diff_bn_okx'] = None
merged_data['perc_diff_bn_bybit'] = None
merged_data['perc_diff_okx_bybit'] = None
merged_data['perc_diff_bn_bitget'] = None   


for index, row in merged_data.iterrows():
    # Extract prices for easier handling
    ex1_price = row['bn_lastPrice']
    ex2_price = row['okx_lastPrice']
    ex3_price = row['bybit_lastPrice']
    ex4_price = row['bitget_lastPrice']
    ex5_price = row['kr_lastPrice']
    ex6_price = row['bingx_lastPrice']
    ex7_price = row['kucoin_lastPrice']

    # Helper function to calculate and store percentage difference
    def calculate_percentage_difference(price1, price2, col_name):
        if pd.notna(price1) and pd.notna(price2) and min(price1, price2) != 0:
            price_diff = abs(price1 - price2)
            percentage_diff = (price_diff / min(price1, price2)) * 100
            merged_data.at[index, col_name] = percentage_diff
        if min(price1, price2) == 0:
            print(f"Debug: Zero price encountered for {price1}, {price2} at row {index}")

    # Calculate percentage differences between all exchange combinations
    calculate_percentage_difference(ex1_price, ex2_price, 'perc_diff_bn_okx')
    calculate_percentage_difference(ex1_price, ex3_price, 'perc_diff_bn_bybit')
    calculate_percentage_difference(ex1_price, ex4_price, 'perc_diff_bn_bitget')
    calculate_percentage_difference(ex1_price, ex5_price, 'perc_diff_bn_kr')
    calculate_percentage_difference(ex1_price, ex6_price, 'perc_diff_bn_bingx')
    calculate_percentage_difference(ex1_price, ex7_price, 'perc_diff_bn_kucoin')

    calculate_percentage_difference(ex2_price, ex3_price, 'perc_diff_okx_bybit')
    calculate_percentage_difference(ex2_price, ex4_price, 'perc_diff_okx_bitget')
    calculate_percentage_difference(ex2_price, ex5_price, 'perc_diff_okx_kr')
    calculate_percentage_difference(ex2_price, ex6_price, 'perc_diff_okx_bingx')
    calculate_percentage_difference(ex2_price, ex7_price, 'perc_diff_okx_kucoin')

    calculate_percentage_difference(ex3_price, ex4_price, 'perc_diff_bybit_bitget')
    calculate_percentage_difference(ex3_price, ex5_price, 'perc_diff_bybit_kr')
    calculate_percentage_difference(ex3_price, ex6_price, 'perc_diff_bybit_bingx')
    calculate_percentage_difference(ex3_price, ex7_price, 'perc_diff_bybit_kucoin')

    calculate_percentage_difference(ex4_price, ex5_price, 'perc_diff_bitget_kr')
    calculate_percentage_difference(ex4_price, ex6_price, 'perc_diff_bitget_bingx')
    calculate_percentage_difference(ex4_price, ex7_price, 'perc_diff_bitget_kucoin')

    calculate_percentage_difference(ex5_price, ex6_price, 'perc_diff_kr_bingx')
    calculate_percentage_difference(ex5_price, ex7_price, 'perc_diff_kr_kucoin')

    calculate_percentage_difference(ex6_price, ex7_price, 'perc_diff_bingx_kucoin')



merged_data.to_excel('futures_price_comparison.xlsx', index=False, engine='openpyxl')


##colouring functions for excel

# Load the workbook and select the active worksheet
wb = load_workbook('futures_price_comparison.xlsx')
ws = wb.active



# Define the color fills for conditional formatting
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red
orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
mint_green_fill = PatternFill(start_color="98FF98", end_color="98FF98", fill_type="solid")  # Mint Green
dark_green_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Dark Green
yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow
white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

# Identify columns for conditional formatting by their header names
columns_to_format = ['perc_diff_bn_okx', 'perc_diff_bn_bybit', 'perc_diff_okx_bybit', 'perc_diff_bn_bitget', 'perc_diff_okx_bitget', 'perc_diff_bybit_bitget', 'perc_diff_bn_kr', 'perc_diff_okx_kr', 'perc_diff_bybit_kr', 'perc_diff_bitget_kr', 'perc_diff_kr_bingx', 'perc_diff_kr_kucoin', 'perc_diff_bingx_kucoin', 'perc_diff_bn_bingx', 'perc_diff_bn_kucoin', 'perc_diff_okx_bingx', 'perc_diff_okx_kucoin', 'perc_diff_bybit_bingx', 'perc_diff_bybit_kucoin', 'perc_diff_bitget_bingx', 'perc_diff_bitget_kucoin']

# Apply conditional formatting
for col in columns_to_format:
    # Get the index of the column (1-based indexing for openpyxl)
    col_idx = merged_data.columns.get_loc(col) + 1
    for row in ws.iter_rows(min_row=2, min_col=col_idx+1, max_col=col_idx+1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Ensure it's a numeric value
                if cell.value < 0.5 * arb_threshold:  # Red
                    cell.fill = red_fill
                elif 0.5 * arb_threshold <= cell.value < 0.75 * arb_threshold:  # Orange
                    cell.fill = orange_fill
                elif 0.75 * arb_threshold <= cell.value < arb_threshold:  # Yellow
                    cell.fill = yellow_fill
                elif arb_threshold <= cell.value < 1.25 * arb_threshold:  # Mint Green
                    cell.fill = mint_green_fill
                elif cell.value >= 1.25 * arb_threshold > 30:  # Dark Green
                    cell.fill = dark_green_fill
                else: 
                    cell.fill = white_fill

                    ##PLAN; MAKE A HASHMAP STRUCTURE WITH SPECTRUM OF COLOURS OF RED TO GREEN, AND THEN ASSIGN THE COLOURS ACCORDING TO THE PERCENTAGE DIFFERENCE. DONE.


# Save the modified workbook
wb.save('futures_price_comparison.xlsx')




#RELATED CODE





# #STAGE2 # for now, this is just a trial, development will be in other side.

# # Rename price columns for each DataFrame

# df_okx = df_okx.rename(columns={'lastPrice': 'OKX Price'})
# df_binance = df_binance.rename(columns={'lastPrice': 'Binance Price'})
# df_bybit = df_bybit.rename(columns={'lastPrice': 'Bybit Price'})
# # Convert price columns to numeric types
# df_okx['OKX Price'] = pd.to_numeric(df_okx['OKX Price'], errors='coerce')
# df_binance['Binance Price'] = pd.to_numeric(df_binance['Binance Price'], errors='coerce')
# df_bybit['Bybit Price'] = pd.to_numeric(df_bybit['Bybit Price'], errors='coerce')

# # Merge the DataFrames
# merged_df = df_okx.merge(df_binance[['symbol', 'Binance Price']], on='symbol', how='outer')
# merged_df = merged_df.merge(df_bybit[['symbol', 'Bybit Price']], on='symbol', how='outer')
# # Calculate price differences
# merged_df['OKX-Binance Diff'] = (merged_df['OKX Price'] - merged_df['Binance Price']).abs()
# merged_df['OKX-Bybit Diff'] = (merged_df['OKX Price'] - merged_df['Bybit Price']).abs()
# merged_df['Binance-Bybit Diff'] = (merged_df['Binance Price'] - merged_df['Bybit Price']).abs()
# # Filter for pairs with significant price differences (e.g., > 1% difference)
# threshold = 0.01  # 1% threshold
# merged_df['OKX-Binance % Diff'] = (merged_df['OKX-Binance Diff'] / merged_df[['OKX Price', 'Binance Price']].min(axis=1)) * 100
# merged_df['OKX-Bybit % Diff'] = (merged_df['OKX-Bybit Diff'] / merged_df[['OKX Price', 'Bybit Price']].min(axis=1)) * 100
# merged_df['Binance-Bybit % Diff'] = (merged_df['Binance-Bybit Diff'] / merged_df[['Binance Price', 'Bybit Price']].min(axis=1)) * 100



# # Filter for rows where any percentage difference exceeds threshold
# significant_diffs = merged_df[(merged_df['OKX-Binance % Diff'] > threshold) |
#                               (merged_df['OKX-Bybit % Diff'] > threshold) |
#                               (merged_df['Binance-Bybit % Diff'] > threshold)]

# # Display the result
# print(merged_df)
# # print(significant_diffs)
# # Save the DataFrame to an Excel file
# # merged_df.to_excel('futures_price_comparison.xlsx', index=False, engine='openpyxl')
# # Save the DataFrame to an Excel file in the current working directory
# merged_df.to_excel('futures_price_comparison.xlsx', index=False, engine='openpyxl')








# #
# #
# #
# #
# #
# # Add a new column to store percentage differences
# merged_data['percentage_diff'] = None

# ##at this stage, we are trying to add percentage diff info to the our merged_data df.

# # List to collect percentage differences for each row in merged_data
# percentage_differences = []


# # Iterate over rows in merged_data
# for index, row in merged_data.iterrows():
#     symbol = row['symbol']
#     prices = row[['bn_lastPrice', 'okx_lastPrice', 'bybit_lastPrice']].dropna()
#     # Convert prices to a list of floats
#     prices = prices.astype(float).values


    
#     # List to collect differences for each row
#     differences = []

#     # Compare prices between every pair of exchanges
#     for i, price1 in enumerate(prices):
#         for j, price2 in enumerate(prices):
#             if i < j:  # Avoid redundant comparisons
#                 # Calculate percentage difference
#                 price_diff = abs(price1 - price2)
#                 percentage_diff = (price_diff / min(price1, price2)) * 100  # Percentage difference calculation

#                 # Check if the percentage difference is 1% or greater
#                 if percentage_diff >= 1:
#                     differences.append(f"{percentage_diff:.2f}% between Exchange {i+1} and Exchange {j+1}")

#     # Join the differences into a single string for the row or set to None if empty
#     percentage_differences.append(", ".join(differences) if differences else None)

# # Add the collected differences to the merged_data DataFrame
# merged_data['percentage_diff'] = percentage_differences
# ##
# #
# #
# #
# #approach for percentage_diff, cancelled


# #DERIBIT DERIBIT DERIBIT DERIBIT
# # Base URL for Deribit's API
# base_url = 'https://www.deribit.com/api/v2/public/'

# # Fetch all instruments for perpetual contracts
# instruments_response = requests.get(base_url + 'get_instruments', params={'kind': 'future'})
# instruments = instruments_response.json()

# # # Filter for perpetual futures
# # perpetuals = [
# #     inst for inst in instruments['result']
# #     if 'PERPETUAL' in inst['instrument_name']
# # ]

# # Filter for perpetual futures
# perpetuals = [
#     inst for inst in instruments['result']
#     if 'PERPETUAL' in inst['instrument_name']
# ]
# # Fetch current ticker information for each perpetual
# data = []
# for inst in perpetuals:
#     symbol = inst['instrument_name']
#     ticker_response = requests.get(base_url + 'ticker', params={'instrument_name': symbol})
#     ticker_data = ticker_response.json()

#     # Extract relevant data
#     last_price = ticker_data['result']['last_price']
#     base_currency = symbol.split('-')[0]  # Extract base currency
#     standardized_symbol = f"{base_currency}USDT"

#     # Append to data list
#     data.append({
#         'symbol': standardized_symbol,
#         'db_lastPrice': last_price
#     })

# # Create a DataFrame
# df_deribit = pd.DataFrame(data)

